"""
EXTRATOR STARLINK - VERSÃO AUTOMÁTICA COM PAGINAÇÃO
Aplica filtro e extrai todos os dados automaticamente
"""
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
import time
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

USER_EMAIL = "fiscal.pulsar@4cta.eb.mil.br"
USER_PASSWORD = "K809(F4a[?"
LOGIN_URL = "https://sport.pulsarconnect.io/login"
STARLINK_URL = "https://sport.pulsarconnect.io/starlink/starlinkMap"

print("="*75)
print(" "*20 + "EXTRATOR STARLINK")
print("="*75)

# Configurar navegador VISÍVEL
options = Options()
options.add_argument("--start-maximized")

driver = webdriver.Chrome(
    service=Service(ChromeDriverManager().install()),
    options=options
)

actions = ActionChains(driver)

print("\n[1/5] Fazendo login...")
driver.get(LOGIN_URL)
time.sleep(5)

email_input = driver.find_element(By.NAME, "userName")
pass_input = driver.find_element(By.NAME, "password")
email_input.send_keys(USER_EMAIL)
pass_input.send_keys(USER_PASSWORD)

login_button = driver.find_element(By.XPATH, "//button[contains(@class, 'loginButton')]")
login_button.click()
time.sleep(15)

print("\n[2/5] Navegando para Starlink...")
driver.get(STARLINK_URL)
time.sleep(15)

print("\n[3/5] Aplicando filtro 'Last 1 Day' automaticamente...")
try:
    # Aguardar página carregar completamente
    time.sleep(8)
    
    # Aguardar algum botão aparecer (indica que a página carregou)
    WebDriverWait(driver, 20).until(
        EC.presence_of_element_located((By.XPATH, "//button"))
    )
    
    # Procurar botão de filtro de data - pode estar com texto "1 Day", "MTD", etc
    date_buttons = driver.find_elements(By.XPATH, "//button")
    
    clicked_filter = False
    for btn in date_buttons:
        btn_text = btn.text.strip()
        if 'Day' in btn_text or 'MTD' in btn_text:
            try:
                # Scroll para o elemento
                driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", btn)
                time.sleep(0.5)
                btn.click()
                print(f"   ✔ Clicou no botão de filtro: '{btn_text}'")
                time.sleep(2)
                clicked_filter = True
                break
            except:
                continue
    
    if clicked_filter:
        # Agora procurar "Last 1 Day" no menu dropdown que abriu
        try:
            # Aguardar o menu aparecer
            time.sleep(1)
            
            # Procurar opção "Last 1 Day"
            last_1_day = driver.find_element(By.XPATH, "//*[text()='Last 1 Day']")
            last_1_day.click()
            print("   ✔ Selecionou 'Last 1 Day'")
            time.sleep(1.5)
            
            # Procurar e clicar botão "Apply"
            apply_btns = driver.find_elements(By.XPATH, "//button[text()='Apply' or contains(text(), 'Apply')]")
            for apply_btn in apply_btns:
                if apply_btn.is_displayed():
                    apply_btn.click()
                    print("   ✔ Clicou em 'Apply'")
                    break
            
            time.sleep(8)  # Aguardar tabela recarregar
            print("   ✔ Filtro 'Last 1 Day' aplicado com sucesso!")
            
        except Exception as e2:
            print(f"   ⚠ Erro ao selecionar 'Last 1 Day': {e2}")
            # Tentar fechar o dropdown pressionando ESC
            actions.send_keys(Keys.ESCAPE).perform()
            time.sleep(2)
    else:
        print("   ⚠ Botão de filtro não encontrado")
        print("   ℹ Continuando com filtro padrão...")
        time.sleep(5)
    
except Exception as e:
    print(f"   ⚠ Erro geral ao aplicar filtro: {e}")
    print("   ℹ Continuando com filtro padrão...")
    time.sleep(5)

print("\n[4/5] Extraindo dados (passando mouse sobre círculos)...")

# Aguardar tabela aparecer
print("   ⏳ Aguardando tabela carregar...")
try:
    WebDriverWait(driver, 30).until(
        EC.presence_of_element_located((By.XPATH, "//table//tbody//tr"))
    )
    time.sleep(5)
except Exception as e:
    print(f"   ❌ Erro ao carregar tabela: {str(e)[:100]}")
    print("   ⚠ Verifique se você está logado e na página correta")
    driver.quit()
    exit(1)

# Aplicar zoom de 75% para melhor visualização
print("   🔍 Aplicando zoom de 75%...")
try:
    driver.execute_script("document.body.style.zoom='75%'")
    time.sleep(2)
except Exception as e:
    print(f"   ⚠ Erro ao aplicar zoom: {str(e)[:50]}")

# Contar total de itens na tabela (todas as páginas)
print("   📊 Contando total de itens na tabela...")
total_items_expected = 0
try:
    # Procurar informação de paginação (ex: "1-25 of 44" ou "1–25 of 44")
    pagination_texts = driver.find_elements(By.XPATH, "//*[contains(text(), 'of') or contains(text(), 'de') or contains(text(), '–')]")
    for text_elem in pagination_texts:
        text = text_elem.text
        # Procurar padrões como "1-25 of 44", "1–25 of 44", etc.
        import re
        # Tentar diferentes padrões de regex
        patterns = [
            r'of\s+(\d+)',  # "of 44"
            r'de\s+(\d+)',  # "de 44"
            r'(\d+)\s*[-–]\s*(\d+)\s+of\s+(\d+)',  # "1-25 of 44"
            r'(\d+)\s*[-–]\s*(\d+)\s+de\s+(\d+)',  # "1-25 de 44"
        ]
        
        for pattern in patterns:
            match = re.search(pattern, text)
            if match:
                # Pegar o último número capturado (que é o total)
                numbers = [int(n) for n in match.groups() if n]
                if numbers:
                    total_items_expected = max(numbers)  # O maior número é geralmente o total
                    print(f"   ✔ Total de itens detectado: {total_items_expected}")
                    break
        
        if total_items_expected > 0:
            break
    
    if total_items_expected == 0:
        print("   ⚠ Não foi possível detectar o total automaticamente")
except Exception as e:
    print(f"   ⚠ Erro ao contar itens: {str(e)[:50]}")

# Scroll para o topo
try:
    driver.execute_script("window.scrollTo(0, 0);")
    time.sleep(2)
except Exception as e:
    print(f"   ❌ Erro ao fazer scroll: {str(e)[:100]}")
    driver.quit()
    exit(1)

all_data = []
kit_ids_processados = set()  # Conjunto para rastrear KIT IDs já processados
identificadores_processados = set()  # Rastrear combinações OM + KIT ID
current_page = 1
has_next_page = True

while has_next_page:
    print(f"\n   📄 Processando página {current_page}...")
    
    # Buscar linhas da tabela
    rows = driver.find_elements(By.XPATH, "//table//tbody//tr[td]")
    
    if not rows or len(rows) == 0:
        print("   ⚠ Tentando seletor alternativo...")
        rows = driver.find_elements(By.XPATH, "//tr[contains(@class, 'MuiTableRow') and .//td]")
    
    print(f"   ✔ {len(rows)} linhas encontradas na página {current_page}\n")
    
    if len(rows) == 0:
        print("   ❌ Nenhuma linha encontrada!")
        break
    
    for idx, row in enumerate(rows):
        try:
            # Extrair células
            cells = row.find_elements(By.TAG_NAME, "td")
            
            if len(cells) < 2:
                continue
            
            # Coluna 1: OM (nome)
            om = cells[0].text.strip()
            
            # Pular cabeçalhos/linhas vazias
            if not om or "SERVICE LINE" in om.upper() or "NO SERVICE" in om.upper() or len(om) < 3:
                continue
            
            # Coluna 2: STATUS
            status_cell = cells[1]
            
            # Buscar SVG circles
            svgs = status_cell.find_elements(By.XPATH, ".//svg | .//*[name()='svg']")
            
            kit_id = ""
            status_cor = "DESCONHECIDO"
            
            # Analisar 2º SVG (índice 1)
            if len(svgs) >= 2:
                segundo_svg = svgs[1]
                
                try:
                    # Scroll suave para centralizar
                    driver.execute_script("arguments[0].scrollIntoView({block: 'center', behavior: 'smooth'});", segundo_svg)
                    time.sleep(0.8)
                    
                    # TENTAR ATÉ 5 VEZES para pegar o tooltip
                    tooltip_found = False
                    max_tentativas = 5
                    
                    for tentativa in range(max_tentativas):
                        if tooltip_found:
                            break
                        
                        # Criar novo ActionChains para cada tentativa (reset)
                        actions = ActionChains(driver)
                        
                        # Mover mouse precisamente sobre o elemento
                        actions.move_to_element(segundo_svg).perform()
                        time.sleep(3.5)  # Tempo para tooltip aparecer
                        
                        # Estratégia 1: Tooltip padrão Material-UI (role='tooltip')
                        try:
                            tooltips = driver.find_elements(By.XPATH, "//div[@role='tooltip']")
                            # Pegar apenas tooltips visíveis e com conteúdo
                            visible_tooltips = [t for t in tooltips if t.is_displayed() and t.size['height'] > 0 and t.size['width'] > 0]
                            
                            # Se houver múltiplos tooltips, pegar o último (mais recente)
                            if visible_tooltips:
                                tooltip = visible_tooltips[-1]
                                tooltip_text = tooltip.text.strip()
                                if tooltip_text and "KIT" in tooltip_text:
                                    words = tooltip_text.replace("\n", " ").split()
                                    for word in words:
                                        if word.startswith("KIT"):
                                            kit_id = word
                                            tooltip_found = True
                                            break
                                    if tooltip_found:
                                        break
                        except:
                            pass
                        
                        # Estratégia 2: Classe com 'tooltip' ou 'Popper'
                        if not tooltip_found:
                            try:
                                tooltips = driver.find_elements(By.XPATH, "//*[contains(@class, 'tooltip') or contains(@class, 'Tooltip') or contains(@class, 'Popper')]")
                                for tooltip in tooltips:
                                    if tooltip.is_displayed() and tooltip.size['height'] > 0:
                                        tooltip_text = tooltip.text.strip()
                                        if tooltip_text and "KIT" in tooltip_text:
                                            words = tooltip_text.replace("\n", " ").split()
                                            for word in words:
                                                if word.startswith("KIT"):
                                                    kit_id = word
                                                    tooltip_found = True
                                                    break
                                            if tooltip_found:
                                                break
                            except:
                                pass
                        
                        # Estratégia 3: Buscar em elementos pai/filhos do SVG
                        if not tooltip_found:
                            try:
                                parent = segundo_svg.find_element(By.XPATH, "..")
                                title_attr = parent.get_attribute('title')
                                aria_label_attr = parent.get_attribute('aria-label')
                                data_tip = parent.get_attribute('data-tip')
                                
                                for attr in [title_attr, aria_label_attr, data_tip]:
                                    if attr and "KIT" in attr:
                                        words = attr.split()
                                        for word in words:
                                            if word.startswith("KIT"):
                                                kit_id = word
                                                tooltip_found = True
                                                break
                                        if tooltip_found:
                                            break
                            except:
                                pass
                        
                        # Estratégia 4: Atributos do próprio SVG
                        if not tooltip_found:
                            title_attr = segundo_svg.get_attribute('title')
                            aria_label_attr = segundo_svg.get_attribute('aria-label')
                            data_tip = segundo_svg.get_attribute('data-tip')
                            
                            for attr in [title_attr, aria_label_attr, data_tip]:
                                if attr and "KIT" in attr:
                                    words = attr.split()
                                    for word in words:
                                        if word.startswith("KIT"):
                                            kit_id = word
                                            tooltip_found = True
                                            break
                                    if tooltip_found:
                                        break
                        
                        # Se não achou ainda, mover mouse fora e tentar de novo
                        if not tooltip_found and tentativa < max_tentativas - 1:
                            # Criar novo ActionChains
                            actions = ActionChains(driver)
                            actions.move_by_offset(200, 0).perform()
                            time.sleep(0.5)
                    
                    # Log se não conseguiu capturar KIT ID após todas as tentativas
                    if not tooltip_found or not kit_id:
                        pass  # Será mostrado no print final como vazio
                    
                    # Análise de COR
                    svg_html = segundo_svg.get_attribute('outerHTML').lower()
                    
                    if 'green' in svg_html or '#00ff00' in svg_html or '#0f0' in svg_html or '#00e676' in svg_html or '#4caf50' in svg_html:
                        status_cor = "VERDE"
                    elif 'red' in svg_html or '#ff0000' in svg_html or '#f00' in svg_html or '#f44336' in svg_html or '#e53935' in svg_html:
                        status_cor = "VERMELHO"
                    
                    # Mover mouse fora
                    actions.move_by_offset(100, 100).perform()
                    time.sleep(0.1)
                    
                except Exception as e:
                    pass
            
            # Coluna 3: USAGE
            usage = cells[2].text.strip() if len(cells) >= 3 else ""
            
            # VERIFICAÇÃO ANTI-DUPLICATA
            # 1. Se o KIT ID já foi usado, tentar capturar novamente (possível erro de hover)
            if kit_id and kit_id in kit_ids_processados:
                print(f"   ⚠ KIT ID duplicado detectado: {kit_id} - Tentando recapturar...")
                
                # Tentar capturar novamente com mais tentativas
                kit_id_backup = kit_id
                kit_id = ""
                
                try:
                    # Reset e nova tentativa de hover mais cuidadosa
                    for tentativa_extra in range(7):  # 7 tentativas extras
                        if kit_id and kit_id != kit_id_backup:
                            break
                        
                        # Mover mouse para longe primeiro
                        actions = ActionChains(driver)
                        actions.move_by_offset(-300, -300).perform()
                        time.sleep(0.5)
                        
                        # Scroll para o elemento
                        driver.execute_script("arguments[0].scrollIntoView({block: 'center', behavior: 'smooth'});", segundo_svg)
                        time.sleep(1)
                        
                        # Novo hover
                        actions = ActionChains(driver)
                        actions.move_to_element(segundo_svg).perform()
                        time.sleep(4)
                        
                        # Tentar capturar tooltip
                        try:
                            tooltips = driver.find_elements(By.XPATH, "//div[@role='tooltip']")
                            visible_tooltips = [t for t in tooltips if t.is_displayed() and t.size['height'] > 0]
                            
                            if visible_tooltips:
                                tooltip = visible_tooltips[-1]
                                tooltip_text = tooltip.text.strip()
                                if tooltip_text and "KIT" in tooltip_text:
                                    words = tooltip_text.replace("\n", " ").split()
                                    for word in words:
                                        if word.startswith("KIT"):
                                            kit_id = word
                                            break
                        except:
                            pass
                
                except Exception as e:
                    pass
                
                # Se ainda é o mesmo KIT ID duplicado, pular
                if not kit_id or kit_id == kit_id_backup:
                    print(f"      ❌ Não foi possível capturar KIT ID diferente - Linha ignorada")
                    continue
                else:
                    print(f"      ✅ KIT ID recapturado com sucesso: {kit_id}")
            
            # 2. Se após recaptura ainda está duplicado, pular
            if kit_id and kit_id in kit_ids_processados:
                print(f"   ⚠ Linha {idx+1} ignorada - KIT ID {kit_id} realmente duplicado")
                continue
            
            # 2. Criar identificador único usando OM + KIT ID
            identificador = f"{om}|{kit_id}" if kit_id else f"{om}|{idx}"
            
            # 3. Verificar se esta combinação exata já foi processada
            if identificador in identificadores_processados:
                print(f"   ⚠ Linha {idx+1} ignorada - Combinação duplicada: {om[:30]} + {kit_id}")
                continue
            
            # 4. Adicionar aos conjuntos de controle
            identificadores_processados.add(identificador)
            if kit_id:
                kit_ids_processados.add(kit_id)
            
            # Adicionar dados
            all_data.append({
                'OM': om,
                'PoP': kit_id if kit_id else "N/A",
                'STATUS': status_cor,
                'OCORRÊNCIA': ''
            })
            
            emoji = "🟢" if status_cor == "VERDE" else ("🔴" if status_cor == "VERMELHO" else "⚪")
            print(f"   {idx+1:2d}. {om[:32]:<32} | {kit_id:<17} | {emoji}")
            
        except Exception as e:
            continue
    
    # Verificar se há próxima página
    try:
        # Procurar botão de próxima página
        next_button = driver.find_element(By.XPATH, "//button[@aria-label='Next page' or contains(@aria-label, 'next') or contains(@class, 'next')]")
        
        # Verificar se o botão está desabilitado
        is_disabled = next_button.get_attribute('disabled')
        
        if is_disabled:
            has_next_page = False
            print(f"\n   ✔ Última página alcançada (página {current_page})")
        else:
            # Clicar no botão de próxima página
            next_button.click()
            print(f"\n   ➡️ Indo para página {current_page + 1}...")
            time.sleep(8)  # Aguardar carregar (tempo aumentado)
            current_page += 1
            
            # Aguardar nova página carregar completamente
            try:
                WebDriverWait(driver, 15).until(
                    EC.presence_of_element_located((By.XPATH, "//table//tbody//tr[td]"))
                )
                time.sleep(3)
            except:
                time.sleep(5)
            
            # Scroll para o topo da nova página
            driver.execute_script("window.scrollTo(0, 0);")
            time.sleep(2)
            
    except:
        # Não encontrou botão de próxima página
        has_next_page = False
        print(f"\n   ✔ Não há mais páginas (total: {current_page} página(s))")

driver.quit()

if not all_data:
    print("\n❌ NENHUM DADO EXTRAÍDO!")
    exit()

print(f"\n✅ {len(all_data)} registros extraídos de {current_page} página(s)!")

# Gerar Excel
print("\n[5/5] Gerando relatório...")

wb = Workbook()
ws = wb.active
ws.title = "Relatório Starlink"

ws.append(['OM', 'PoP', 'STATUS', 'OCORRÊNCIA'])

# Cabeçalho
header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
for cell in ws[1]:
    cell.fill = header_fill
    cell.font = Font(bold=True, color="FFFFFF")
    cell.alignment = Alignment(horizontal="center", vertical="center")

# Bordas
border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                top=Side(style='thin'), bottom=Side(style='thin'))

# Dados
for item in all_data:
    ws.append([item['OM'], item['PoP'], item['STATUS'], item['OCORRÊNCIA']])

# Cores
verde = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
vermelho = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
amarelo = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

for row in ws.iter_rows(min_row=2, max_col=4):
    for cell in row:
        cell.border = border
    
    status_cell = row[2]
    if status_cell.value == "VERDE":
        status_cell.fill = verde
        status_cell.font = Font(bold=True)
    elif status_cell.value == "VERMELHO":
        status_cell.fill = vermelho
        status_cell.font = Font(color="FFFFFF", bold=True)
    else:
        status_cell.fill = amarelo

# Larguras
ws.column_dimensions['A'].width = 45
ws.column_dimensions['B'].width = 18
ws.column_dimensions['C'].width = 15
ws.column_dimensions['D'].width = 30

wb.save("Relatorio_Starlink_Final.xlsx")
df = pd.DataFrame(all_data)
df.to_csv("Relatorio_Starlink_Final.csv", index=False, encoding='utf-8-sig')

print("\n" + "="*75)
print("✅ CONCLUÍDO!")
print("="*75)
print(f"📊 Total: {len(all_data)} registros de {current_page} página(s)")
print(f"   🟢 Verde: {len([x for x in all_data if x['STATUS'] == 'VERDE'])}")
print(f"   🔴 Vermelho: {len([x for x in all_data if x['STATUS'] == 'VERMELHO'])}")
print(f"   ⚪ Desconhecido: {len([x for x in all_data if x['STATUS'] == 'DESCONHECIDO'])}")
print(f"   📋 KIT IDs capturados: {len([x for x in all_data if x['PoP'] != 'N/A'])}")

# Comparar com total esperado (se foi detectado corretamente)
if total_items_expected > 0 and total_items_expected >= 10:  # Ignorar se o número for muito pequeno (erro de detecção)
    diferenca = total_items_expected - len(all_data)
    if diferenca > 0:
        print(f"\n⚠ ATENÇÃO: {diferenca} item(s) não foram capturados!")
        print(f"   Esperado: {total_items_expected} | Capturado: {len(all_data)}")
    elif diferenca < 0:
        print(f"\n⚠ ATENÇÃO: {abs(diferenca)} item(s) a mais foram capturados!")
        print(f"   Esperado: {total_items_expected} | Capturado: {len(all_data)}")
    else:
        print(f"\n✅ Todos os {total_items_expected} itens foram capturados com sucesso!")

print(f"\n📁 Arquivos:")
print(f"   • Relatorio_Starlink_Final.xlsx")
print(f"   • Relatorio_Starlink_Final.csv")
print("="*75)
