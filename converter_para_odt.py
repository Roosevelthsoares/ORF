"""
Conversor Excel para ODT
Converte o arquivo Relatorio_Starlink_Final.xlsx para formato ODT
mantendo apenas as cores das células (sem texto de cor)
"""
import pandas as pd
from openpyxl import load_workbook
from odf.opendocument import OpenDocumentText
from odf.style import Style, TableCellProperties, TextProperties, TableColumnProperties, TableProperties
from odf.table import Table, TableRow, TableCell, TableColumn
from odf.text import P

print("="*75)
print(" "*20 + "CONVERSOR EXCEL → ODT")
print("="*75)

# Dicionário de mapeamento KIT ID -> Nome da OM
MAPEAMENTO_OM = {
    "KIT304062259": "Cmdo 1ª Bda Inf Sl",
    "KIT304059560": "Cmdo 2ª Bda Inf Sl",
    "KIT304135659": "2ª Bda Inf Sl",
    "KITP00237489": "Cmdo 16ª Bda Inf Sl",
    "KIT304132110": "Cmdo 17ª Bda Inf Sl",
    "KIT304059859": "4º BIS - DEF - Epitaciolândia",
    "KIT304039763": "4º BIS - 2º PEF - Assis Brasil",
    "KIT304131574": "4º BIS - 3º PEF - Plácido de Castro",
    "KIT304039768": "4º BIS - 4º PEF - Santa Rosa do Purus",
    "KIT3040397634": "5º BIS – 1º PEF - Yauaretê",
    "KIT304039771": "5º BIS – 2º PEF - Querari",
    "KIT304039241": "5º BIS – 3º PEF - São Joaquim",
    "KIT303910747": "5º BIS – 4º PEF - Cucuí",
    "KIT304039236": "5º BIS – 5º PEF - Maturacá",
    "KIT304039230": "5º BIS – 6º PEF - Pari-Cachoeira",
    "KIT304039765": "5º BIS – 7º PEF - Tunuí",
    "KIT304135657": "7º BIS - 1º PEF - Bonfim",
    "KIT304059865": "7º BIS - 2º PEF - Normandia",
    "KIT304059878": "7º BIS - 3º PEF - Pacaraima",
    "KIT304039242": "7º BIS - 4º PEF - Surucucu",
    "KIT304039235": "7º BIS - 5º PEF - Auaris",
    "KIT304044880": "7º BIS - 6º PEF - Uiramutã",
    "KIT304059852": "7º BIS - Base Pakilapi",
    "KIT304059547": "7º BIS - Base Kaianaú",
    "KIT303901850": "7º BIS - DEF Waikas",
    "KIT303904970": "8º BIS – 1º PEF-Palmeira do Javari",
    "KIT304059879": "8º BIS - 2º PEF - Ipiranga",
    "KIT304059946": "8º BIS - 4º PEF - Estirão do Equador",
    "KIT304039752": "61º BIS - DEF- Marechal Thaumaturgo",
    "KIT304132549": "34º BIS - Oiapoque",
    "KIT303903287": "34º BIS - Vila Brasil",
    "KIT304131555": "34º BIS - Tiriós",
    "KIT304039747": "3º BIS",
    "KIT303909856": "6º BIS",
    "KIT304132264": "6º BIS - 1º PEF - Príncipe da Beira",
    "KIT303844328": "17º BIS",
    "KIT303847361": "17º BIS – 3º PEF-Vila Bittencourt",
    "KIT304039751": "HGuT",
    "KIT304059544": "2º B Log Sl",
    "KIT304039748": "21ª Cia E Cnst",
    "KIT304132127": "7º BEC (Destacamento)",
    "KIT304145670": "BI-02(CIGS)",
    "KIT303729090": "CMDO 8º BIS - Tabatinga",
    "KIT304132551": "4º CTA 02 - Manaus",
    "KIT304145658": "Cmdo 6º BIS 02",
    "KIT304132540": "2º PEF - Normandia",
    "KIT304132552": "3º PEF - Vila Bittencourt",
    "KIT304132336": "Querari - AM/CFRN",
    "KIT304145662": "4º CTA 01 - Manaus",
    "KIT304059853": "1º PEF Yauaretê"
}

# Carregar arquivo Excel
print("\n[1/3] Lendo arquivo Excel...")
try:
    wb = load_workbook("Relatorio_Starlink_Final.xlsx")
    ws = wb.active
    print(f"   ✔ Arquivo carregado: {ws.max_row} linhas")
except FileNotFoundError:
    print("   ❌ Erro: Arquivo 'Relatorio_Starlink_Final.xlsx' não encontrado!")
    print("   Execute primeiro o script 'extrair_relatorio_final.py'")
    exit()

# Criar documento ODT de texto
print("\n[2/3] Criando documento ODT...")
doc = OpenDocumentText()

# Definir estilo de borda para todas as células
border_style = "0.05pt solid #000000"

# Criar estilos para as cores com borda
# Verde (#00B050)
style_verde = Style(name="verde", family="table-cell")
style_verde.addElement(TableCellProperties(backgroundcolor="#00B050", border=border_style))
doc.automaticstyles.addElement(style_verde)

# Vermelho (#FF0000)
style_vermelho = Style(name="vermelho", family="table-cell")
style_vermelho.addElement(TableCellProperties(backgroundcolor="#FF0000", border=border_style))
doc.automaticstyles.addElement(style_vermelho)

# Amarelo (#FFFF00)
style_amarelo = Style(name="amarelo", family="table-cell")
style_amarelo.addElement(TableCellProperties(backgroundcolor="#FFFF00", border=border_style))
doc.automaticstyles.addElement(style_amarelo)

# Cabeçalho (azul #366092 com texto branco)
style_header = Style(name="header", family="table-cell")
style_header.addElement(TableCellProperties(backgroundcolor="#366092", border=border_style))
style_header.addElement(TextProperties(color="#FFFFFF", fontweight="bold"))
doc.automaticstyles.addElement(style_header)

# Estilo normal (células sem cor, mas com borda)
style_normal = Style(name="normal", family="table-cell")
style_normal.addElement(TableCellProperties(border=border_style))
doc.automaticstyles.addElement(style_normal)

# Estilo para a tabela
table_style = Style(name="table_style", family="table")
table_style.addElement(TableProperties(width="16cm", align="left"))
doc.automaticstyles.addElement(table_style)

# Estilos para as colunas com larguras específicas
col_style_om = Style(name="col_om", family="table-column")
col_style_om.addElement(TableColumnProperties(columnwidth="7cm"))
doc.automaticstyles.addElement(col_style_om)

col_style_pop = Style(name="col_pop", family="table-column")
col_style_pop.addElement(TableColumnProperties(columnwidth="4cm"))
doc.automaticstyles.addElement(col_style_pop)

col_style_status = Style(name="col_status", family="table-column")
col_style_status.addElement(TableColumnProperties(columnwidth="3cm"))
doc.automaticstyles.addElement(col_style_status)

col_style_ocorrencia = Style(name="col_ocorrencia", family="table-column")
col_style_ocorrencia.addElement(TableColumnProperties(columnwidth="2cm"))
doc.automaticstyles.addElement(col_style_ocorrencia)

# Criar tabela
table = Table(name="Relatório Starlink", stylename=table_style)

# Adicionar colunas à tabela com larguras específicas
table.addElement(TableColumn(stylename=col_style_om))
table.addElement(TableColumn(stylename=col_style_pop))
table.addElement(TableColumn(stylename=col_style_status))
table.addElement(TableColumn(stylename=col_style_ocorrencia))

# Processar todas as linhas do Excel e ordenar por OM
print(f"   ✔ Processando {ws.max_row} linhas...")

# Coletar todos os dados em uma lista
dados = []
cabecalho = None

for excel_row in ws.iter_rows(min_row=1, max_row=ws.max_row):
    if excel_row[0].row == 1:
        # Guardar cabeçalho
        cabecalho = excel_row
    else:
        # Coletar dados das linhas
        kit_id = str(excel_row[1].value) if excel_row[1].value else ""
        om_nome = MAPEAMENTO_OM.get(kit_id, str(excel_row[0].value) if excel_row[0].value else "")
        dados.append((om_nome, excel_row))

# Ordenar dados por nome da OM (alfabeticamente)
dados.sort(key=lambda x: x[0])

# Processar cabeçalho primeiro
odt_row = TableRow()
for excel_cell in cabecalho:
    cell_value = str(excel_cell.value) if excel_cell.value else ""
    cell = TableCell(stylename=style_header)
    cell.addElement(P(text=cell_value))
    odt_row.addElement(cell)
table.addElement(odt_row)

rows_processed = 1

# Processar linhas ordenadas
for om_nome, excel_row in dados:
    odt_row = TableRow()
    
    # Guardar o KIT ID da linha para fazer o mapeamento
    kit_id = str(excel_row[1].value) if excel_row[1].value else ""
    
    for idx, excel_cell in enumerate(excel_row):
        cell_value = str(excel_cell.value) if excel_cell.value else ""
        
        # Coluna STATUS (índice 2) - aplicar cor baseada no preenchimento do Excel
        if idx == 2:
            # Pegar cor de preenchimento do Excel
            fill_color = excel_cell.fill.start_color.rgb if excel_cell.fill.start_color else None
            
            # Converter cor do Excel para estilo ODT
            if fill_color:
                # Remover alpha channel se existir (ex: FF00B050 -> 00B050)
                if len(fill_color) == 8:
                    fill_color = fill_color[2:]
                
                fill_color = fill_color.upper()
                
                # Verde
                if fill_color in ['00B050', '00FF00', '0F0']:
                    cell = TableCell(stylename=style_verde)
                # Vermelho
                elif fill_color in ['FF0000', 'F00']:
                    cell = TableCell(stylename=style_vermelho)
                # Amarelo
                elif fill_color in ['FFFF00', 'FF0']:
                    cell = TableCell(stylename=style_amarelo)
                else:
                    cell = TableCell(stylename=style_normal)
            else:
                cell = TableCell(stylename=style_normal)
            
            # Célula STATUS fica vazia (sem texto)
            cell.addElement(P(text=''))
        else:
            # Coluna OM (índice 0) - substituir pelo nome correto usando KIT ID
            if idx == 0 and kit_id in MAPEAMENTO_OM:
                cell = TableCell(stylename=style_normal)
                cell.addElement(P(text=MAPEAMENTO_OM[kit_id]))
            else:
                # Outras colunas: texto normal com borda
                cell = TableCell(stylename=style_normal)
                cell.addElement(P(text=cell_value))
        
        odt_row.addElement(cell)
    
    table.addElement(odt_row)
    rows_processed += 1

doc.text.addElement(table)

# Salvar arquivo ODT
print("\n[3/3] Salvando arquivo ODT...")
try:
    doc.save("Relatorio_Starlink_Final.odt")
    print(f"   ✔ {rows_processed} linhas convertidas")
    print(f"   ✔ Arquivo salvo: Relatorio_Starlink_Final.odt")
except Exception as e:
    print(f"   ❌ Erro ao salvar: {e}")
    exit()

print("\n" + "="*75)
print("✅ CONVERSÃO CONCLUÍDA!")
print("="*75)
print("\nDetalhes:")
print(f"   📄 Linhas processadas: {rows_processed}")
print(f"   🎨 Cores preservadas:")
print(f"      • Verde (#00B050) - Status OK")
print(f"      • Vermelho (#FF0000) - Status com problema")
print(f"      • Amarelo (#FFFF00) - Status desconhecido")
print(f"\n   📁 Arquivo gerado: Relatorio_Starlink_Final.odt")
print(f"   📂 Compatível com: LibreOffice, OpenOffice")
print("="*75)
