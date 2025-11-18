"""
Script para gerar relatório de status do Starlink
Formato: OM | PoP | STATUS | OCORRÊNCIA
"""
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
import time
import json
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

# ============================================================
# CONFIGURAÇÕES
# ============================================================
USER_EMAIL = "fiscal.pulsar@4cta.eb.mil.br"
USER_PASSWORD = "K809(F4a[?"
LOGIN_URL = "https://sport.pulsarconnect.io/login"

# ============================================================
# FUNÇÃO PARA FAZER LOGIN E OBTER TOKEN
# ============================================================
def fazer_login():
    options = Options()
    # Descomente para modo invisível:
    # options.add_argument("--headless=new")
    options.add_argument("--disable-gpu")
    options.add_argument("--window-size=1920,1080")
    
    driver = webdriver.Chrome(
        service=Service(ChromeDriverManager().install()),
        options=options
    )
    
    print("➡️ Fazendo login...")
    driver.get(LOGIN_URL)
    
    wait = WebDriverWait(driver, 20)
    time.sleep(5)
    
    try:
        email_input = wait.until(EC.presence_of_element_located((By.NAME, "userName")))
        pass_input = wait.until(EC.presence_of_element_located((By.NAME, "password")))
        
        email_input.send_keys(USER_EMAIL)
        pass_input.send_keys(USER_PASSWORD)
        
        login_button = driver.find_element(By.XPATH, "//button[contains(@class, 'loginButton')]")
        login_button.click()
        
        print("⏳ Aguardando login...")
        time.sleep(12)
        
        # Obter token do localStorage
        for tentativa in range(10):
            logged_in_user = driver.execute_script("return window.localStorage.getItem('loggedInUser');")
            if logged_in_user:
                logged_data = json.loads(logged_in_user)
                token = logged_data["data"]["access_token"]
                print("✔ Token obtido!")
                driver.quit()
                return token
            time.sleep(2)
        
        driver.quit()
        return None
        
    except Exception as e:
        print(f"❌ Erro no login: {e}")
        driver.quit()
        return None

# ============================================================
# FUNÇÃO PARA BUSCAR SERVICE LINES
# ============================================================
def buscar_service_lines(token):
    import requests
    
    headers = {"Authorization": f"Bearer {token}"}
    
    # API para listar service lines
    url = "https://api.k4mobility.com/iam/starlink/serviceLine/dpId/DP-0833"
    
    print("\n➡️ Buscando service lines...")
    response = requests.get(url, headers=headers)
    
    if response.status_code == 200:
        data = response.json()
        if 'serviceLines' in data:
            print(f"✔ {len(data['serviceLines'])} service lines encontradas")
            return data['serviceLines']
    
    return []

# ============================================================
# FUNÇÃO PARA BUSCAR STATUS DETALHADO
# ============================================================
def buscar_status_detalhado(token, service_line_id):
    import requests
    
    headers = {"Authorization": f"Bearer {token}"}
    
    # Tentar diferentes endpoints para obter status
    urls = [
        f"https://api.k4mobility.com/starlink/serviceLine/{service_line_id}/status",
        f"https://api.k4mobility.com/starlink/serviceLine/{service_line_id}/health",
        f"https://api.k4mobility.com/starlink/serviceLine/{service_line_id}",
    ]
    
    for url in urls:
        try:
            response = requests.get(url, headers=headers, timeout=5)
            if response.status_code == 200:
                return response.json()
        except:
            continue
    
    return None

# ============================================================
# PROCESSAR DADOS
# ============================================================
token = fazer_login()

if not token:
    print("❌ Não foi possível obter o token")
    exit()

service_lines = buscar_service_lines(token)

if not service_lines:
    print("❌ Nenhuma service line encontrada")
    exit()

print("\n➡️ Processando dados...")

relatorio = []

for idx, sl in enumerate(service_lines):
    om = sl.get('name', '')
    pop = sl.get('id', '')  # O ID pode ser o KIT
    
    # Se o ID não tem formato KIT, procurar em outros campos
    if not pop.startswith('KIT'):
        pop = sl.get('userTerminalId', '') or sl.get('kitId', '') or pop
    
    # Tentar determinar o status
    # Por padrão, vamos marcar como "DESCONHECIDO" e você pode ajustar manualmente
    status = "DESCONHECIDO"
    
    # Se houver informações de status na service line
    if 'status' in sl:
        status_info = sl.get('status', '')
        if 'online' in str(status_info).lower() or 'active' in str(status_info).lower():
            status = "VERDE"
        elif 'offline' in str(status_info).lower() or 'inactive' in str(status_info).lower():
            status = "VERMELHO"
    
    relatorio.append({
        'OM': om,
        'PoP': pop,
        'STATUS': status,
        'OCORRÊNCIA': ''
    })
    
    print(f"  {idx+1}/{len(service_lines)}: {om[:40]}")

print(f"\n✔ {len(relatorio)} registros processados!")

# ============================================================
# GERAR EXCEL FORMATADO
# ============================================================
print("\n➡️ Gerando planilha Excel...")

wb = Workbook()
ws = wb.active
ws.title = "Relatório Starlink"

# Cabeçalhos
headers = ['OM', 'PoP', 'STATUS', 'OCORRÊNCIA']
ws.append(headers)

# Formatação do cabeçalho
header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=12)
header_alignment = Alignment(horizontal="center", vertical="center")

for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = header_alignment

# Bordas
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Adicionar dados
for item in relatorio:
    ws.append([item['OM'], item['PoP'], item['STATUS'], item['OCORRÊNCIA']])

# Formatação das células
verde_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
vermelho_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
amarelo_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=4):
    # Aplicar bordas
    for cell in row:
        cell.border = thin_border
        cell.alignment = Alignment(vertical="center")
    
    # Colorir coluna STATUS baseado no valor
    status_cell = row[2]  # Coluna STATUS (índice 2)
    if status_cell.value == "VERDE":
        status_cell.fill = verde_fill
    elif status_cell.value == "VERMELHO":
        status_cell.fill = vermelho_fill
    else:
        status_cell.fill = amarelo_fill

# Ajustar largura das colunas
ws.column_dimensions['A'].width = 40  # OM
ws.column_dimensions['B'].width = 20  # PoP
ws.column_dimensions['C'].width = 15  # STATUS
ws.column_dimensions['D'].width = 30  # OCORRÊNCIA

# Salvar
wb.save("Relatorio_Starlink.xlsx")
print("✔ Planilha salva: Relatorio_Starlink.xlsx")

# Também salvar em CSV simples
df = pd.DataFrame(relatorio)
df.to_csv("Relatorio_Starlink.csv", index=False, encoding='utf-8-sig')
print("✔ CSV salvo: Relatorio_Starlink.csv")

print("\n✅ Concluído!")
print(f"📊 Total de registros: {len(relatorio)}")
print(f"   - Verde: {len([r for r in relatorio if r['STATUS'] == 'VERDE'])}")
print(f"   - Vermelho: {len([r for r in relatorio if r['STATUS'] == 'VERMELHO'])}")
print(f"   - Desconhecido: {len([r for r in relatorio if r['STATUS'] == 'DESCONHECIDO'])}")
